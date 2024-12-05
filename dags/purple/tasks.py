import base64
from datetime import datetime
from io import BytesIO
import logging
from airflow.decorators import task, dag
from airflow.models import Variable
import pandas as pd
from utils.google_cloud import get_bq_client, get_gmail_client
from openpyxl import load_workbook
import json

DF_COLS = ["date", "brand_name", "category_name", "ean_code", "sku", "nmv", "mrp", "qty", 
           "orders", "price_off_per", "abs_price_off", "off_invoice"]

TABLE_NAME = "pilgrim_bi_purple.claims_report"

@dag("purple_email_claims", schedule='0 22 * * *', start_date=datetime(year=2024,month=12,day=5), tags=["purple"])
def purple_email_claims():

    @task.python
    def fetch_and_load_mails():
        logging.info(f"Started Task")
        credentials_info = Variable.get("GOOGLE_BIGQUERY_CREDENTIALS")
        GMAIL_FB_APP_TESTING_TOKEN = json.loads(Variable.get("GMAIL_FB_APP_TESTING_TOKEN", "{}"))
        gmail_client = get_gmail_client(GMAIL_FB_APP_TESTING_TOKEN)
        bq_client = get_bq_client(credentials_info)
        messages_client = gmail_client.users().messages()
        attachments_client = messages_client.attachments()

        mails = []
        get_mails_req = messages_client.list(userId="me", q="subject:pilgrim claims AND has:attachment AND is:unread", maxResults=10)
        get_mails_resp = get_mails_req.execute()

        while( isinstance(get_mails_resp, (dict,)) and get_mails_resp.get("nextPageToken") is not None):
            mails.extend(get_mails_resp['messages'])
            get_mails_resp = messages_client.list_next(get_mails_req, get_mails_resp).execute()

        mails.extend(get_mails_resp.get('messages', []))

        for mail in mails:
            email_data = messages_client.get(userId="me", id=mail["id"]).execute()
            mail_attachment = list(filter(lambda x:x['filename'].lower().find("xlsx")>=0, email_data['payload']['parts']))[0]
            logging.info(f"processing {mail_attachment['filename']}")
            attachment_data = attachments_client.get(userId="me", messageId=mail["id"], id=mail_attachment['body']['attachmentId']).execute()
            attachment_bytes = BytesIO(base64.urlsafe_b64decode(attachment_data['data'] + '=' * (4 - len(attachment_data['data']) % 4)))
            
            wb = load_workbook(attachment_bytes, read_only=True)
            sheet = wb[wb.sheetnames[0]]

            i = 0
            for x in sheet.iter_rows():
                if x[0].value is not None: break
                i += 1

            df = pd.read_excel(attachment_bytes, header=i, usecols='A:L', names=DF_COLS)
            df["runid"] = int(datetime.now().strftime("%Y%m%d%H%M%S"))
            df['date'] = df['date'].dt.date
            table = bq_client.get_table(TABLE_NAME)
            bq_client.insert_rows_from_dataframe(TABLE_NAME, df, selected_fields=table.schema)
            messages_client.modify(userId="me", id=mail["id"], body=dict(removeLabelIds=["UNREAD"])).execute()
    
    resp = fetch_and_load_mails()

purple_email_claims_dag = purple_email_claims()