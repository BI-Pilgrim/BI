import json
import logging
import os.path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from airflow.models import Variable
from utils.google_cloud import get_bq_client, get_gmail_client
from io import BytesIO
import base64
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# If modifying these scopes, delete the file token.json.
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

GMAIL_FB_APP_TESTING_TOKEN = json.loads(Variable.get("GMAIL_FB_APP_TESTING_TOKEN", "{}"))

creds = Credentials.from_authorized_user_info(GMAIL_FB_APP_TESTING_TOKEN)
creds.refresh(Request())
service = build("gmail", "v1", credentials=creds)
results = service.users().labels().list(userId="me").execute()

messages_client = service.users().messages()
attachments_client = messages_client.attachments()


DF_COLS = ["date", "brand_name", "category_name", "ean_code", "sku", "nmv", "mrp", "qty", 
           "orders", "price_off_per", "abs_price_off", "off_invoice"]

TABLE_NAME = "pilgrim_bi_purple.claims_report"

credentials_info = Variable.get("GOOGLE_BIGQUERY_CREDENTIALS")
GMAIL_FB_APP_TESTING_TOKEN = json.loads(Variable.get("GMAIL_FB_APP_TESTING_TOKEN", "{}"))
gmail_client = get_gmail_client(GMAIL_FB_APP_TESTING_TOKEN)
bq_client = get_bq_client(credentials_info)

mails = []
get_mails_req = messages_client.list(userId="me", q="subject:pilgrim claims AND has:attachment AND is:unread", maxResults=10)
get_mails_resp = get_mails_req.execute()

while( isinstance(get_mails_resp, (dict,)) and get_mails_resp.get("nextPageToken") is not None):
    mails.extend(get_mails_resp['messages'])
    get_mails_resp = messages_client.list_next(get_mails_req, get_mails_resp).execute()

mails.extend(get_mails_resp.get('messages', []))

for mail in mails:
    email_data = messages_client.get(userId="me", id=mail["id"]).execute()
    mail_attachment = list(filter(lambda x:x['filename'].lower().find("xlsx")>=0, email_data['payload']['parts']))[0]
    logging.info(f"processing {mail_attachment['filename']}")
    attachment_data = attachments_client.get(userId="me", messageId=mail["id"], id=mail_attachment['body']['attachmentId']).execute()
    attachment_bytes = BytesIO(base64.urlsafe_b64decode(attachment_data['data'] + '=' * (4 - len(attachment_data['data']) % 4)))
    
    wb = load_workbook(attachment_bytes, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    
    i = 0
    for x in sheet.iter_rows():
      if x[0].value is not None: break
      i += 1

    df = pd.read_excel(attachment_bytes, header=i, usecols='A:L', names=DF_COLS)
    df['date'] = df['date'].dt.date
    df["runid"] = int(datetime.now().strftime("%Y%m%d%H%M%S"))
    table = bq_client.get_table(TABLE_NAME)
    bq_client.insert_rows_from_dataframe(TABLE_NAME, df, selected_fields=table.schema)
    messages_client.modify(userId="me", id=mail["id"], body=dict(removeLabelIds=["UNREAD"])).execute()


def main():
  """Shows basic usage of the Gmail API.
  Lists the user's Gmail labels.
  """
  creds = None
  # The file token.json stores the user's access and refresh tokens, and is
  # created automatically when the authorization flow completes for the first
  # time.
  if os.path.exists("token.json"):
    creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    
  # If there are no (valid) credentials available, let the user log in.
  if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
      creds.refresh(Request())
    else:
      flow = InstalledAppFlow.from_client_secrets_file(
          "credentials.json", SCOPES
      )
      creds = flow.run_local_server(port=0)
    # Save the credentials for the next run
    with open("token.json", "w") as token:
      token.write(creds.to_json())

  try:
    # Call the Gmail API
    service = build("gmail", "v1", credentials=creds)
    results = service.users().labels().list(userId="me").execute()
    labels = results.get("labels", [])

    if not labels:
      print("No labels found.")
      return
    print("Labels:")
    for label in labels:
      print(label["name"])

  except HttpError as error:
    # TODO(developer) - Handle errors from gmail API.
    print(f"An error occurred: {error}")


if __name__ == "__main__":
  main()